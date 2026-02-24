import requests
from bs4 import BeautifulSoup
import time
import json
import os
import logging
from datetime import datetime
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from telegram import Update
from telegram.ext import Updater, CommandHandler, CallbackContext

# ЗАГРУЗКА .env
load_dotenv()

TOKEN = os.getenv("TOKEN")
CHAT_ID = os.getenv("CHAT_ID")

# НАСТРОЙКИ
CHECK_INTERVAL = 3600
DATA_FILE = "prices.json"
EXCEL_FILE = "prices.xlsx"

PRODUCTS = {
    "iphone": {
        "url": "https://www.wildberries.ru/catalog/574048913/detail.aspx?size=785912405",
        "target_price": 50000
    },
}
#

logging.basicConfig(level=logging.INFO)

# Работа с JSON
def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, "r") as f:
            return json.load(f)
    return {}

def save_data(data):
    with open(DATA_FILE, "w") as f:
        json.dump(data, f, indent=4)

# Excel
def save_to_excel(name, price):
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Товар", "Цена", "Дата"])
        wb.save(EXCEL_FILE)

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([name, price, datetime.now().strftime("%d-%m-%Y %H:%M")])
    wb.save(EXCEL_FILE)

# График
def generate_chart(name):
    data = load_data()
    prices = data.get(name, {}).get("history", [])

    if not prices:
        return None

    values = [p["price"] for p in prices]
    dates = [p["date"] for p in prices]

    plt.figure()
    plt.plot(dates, values)
    plt.xticks(rotation=45)
    plt.title(f"Изменение цены: {name}")
    plt.tight_layout()

    filename = f"{name}_chart.png"
    plt.savefig(filename)
    plt.close()
    return filename

# Парсинг
def parse_price(url):
    try:
        # Извлекаем артикул из ссылки (цифры в середине)
        # Пример: https://www.wildberries.ru/catalog/12345678/detail.aspx
        item_id = ''.join(filter(str.isdigit, url.split('/')[-2]))
        
        # Обращаемся к официальному API Wildberries для карточек товара
        api_url = f"https://card.wb.ru/cards/v1/detail?appType=1&curr=rub&dest=-1257786&nm={item_id}"
        
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36"
        }
        
        response = requests.get(api_url, headers=headers, timeout=10)
        data = response.json()
        
        # Получаем цену (она приходит в копейках, поэтому делим на 100)
        price = data['data']['products'][0]['salePriceU'] / 100
        return int(price)
        
    except Exception as e:
        logging.error(f"Ошибка парсинга Wildberries: {e}")
        return None
    
# Мониторинг
def check_prices(context: CallbackContext):
    data = load_data()

    for name, info in PRODUCTS.items():
        current_price = parse_price(info["url"])

        if current_price is None:
            continue

        if name not in data:
            data[name] = {"history": [], "last_price": None}

        last_price = data[name]["last_price"]

        if current_price != last_price:
            data[name]["last_price"] = current_price
            data[name]["history"].append({
                "price": current_price,
                "date": datetime.now().strftime("%d-%m %H:%M")
            })

            save_to_excel(name, current_price)

            if current_price <= info["target_price"]:
                context.bot.send_message(
                    chat_id=context.job.context,
                    text=f"🔥 {name} упал до {current_price} ₽"
                )

    save_data(data)

# Команды бота
def start(update: Update, context: CallbackContext):
    update.message.reply_text(
        "🤖 Бот мониторинга цен\n"
        "/price - текущие цены\n"
        "/chart - график\n"
        "/excel - выгрузить Excel"
    )

def price(update: Update, context: CallbackContext):
    data = load_data()
    message = ""

    for name in data:
        message += f"{name}: {data[name]['last_price']} ₽\n"

    update.message.reply_text(message or "Нет данных")

def chart(update: Update, context: CallbackContext):
    for name in PRODUCTS:
        file = generate_chart(name)
        if file:
            update.message.reply_photo(photo=open(file, "rb"))

def excel(update: Update, context: CallbackContext):
    if os.path.exists(EXCEL_FILE):
        update.message.reply_document(document=open(EXCEL_FILE, "rb"))

# Запуск
def main():
    updater = Updater(TOKEN, use_context=True)
    dp = updater.dispatcher

    dp.add_handler(CommandHandler("start", start))
    dp.add_handler(CommandHandler("price", price))
    dp.add_handler(CommandHandler("chart", chart))
    dp.add_handler(CommandHandler("excel", excel))

    updater.start_polling()

    #автопроверка
    chat_id = "ТВОЙ_CHAT_ID"
    updater.job_queue.run_repeating(
        check_prices,
        interval=CHECK_INTERVAL,
        first=10,
        context=chat_id
    )

    updater.idle()

if __name__ == "__main__":
    main()

